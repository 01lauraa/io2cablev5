"""Regenerate build_config.py from the live config/kabelconfig.xlsx.

Run this after editing the workbook, so the bootstrap script never drifts out of
sync with reality:

    python3 dump_config.py

Drift is what made the previous build_config.py dangerous: the workbook had
gained three parameters, per-family columns and the 6_Locatiekoppen tab, while
the script still wrote the old version -- so re-running it would silently destroy
config. Keep them in step.
"""
from openpyxl import load_workbook

HEADER = '''"""Bootstrap a fresh config/kabelconfig.xlsx.

AUTO-GENERATED from the live workbook by dump_config.py -- do not hand-edit.

The workbook is the source of truth and the estimators edit it directly. This
script exists only to recreate it from scratch (new install, or recovering a
corrupted file). Running it OVERWRITES config/kabelconfig.xlsx -- take a copy of
any project parameters first.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

HDR = Font(bold=True, color="FFFFFF", name="Arial", size=10)
FILL = PatternFill("solid", start_color="1F4E78")
BODY = Font(name="Arial", size=10)
NOTE = Font(name="Arial", size=9, italic=True, color="666666")


def sheet(wb, name, note, headers, rows, widths):
    ws = wb.create_sheet(name)
    if note:
        ws.append([note]); ws.cell(1, 1).font = NOTE
        ws.append([])
    ws.append(headers)
    hr = ws.max_row
    for c in range(1, len(headers) + 1):
        cell = ws.cell(hr, c)
        cell.font, cell.fill = HDR, FILL
        cell.alignment = Alignment(vertical="center")
    for r in rows:
        ws.append(list(r))
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    for row in ws.iter_rows(min_row=hr + 1):
        for c in row:
            c.font = BODY
    ws.freeze_panes = ws.cell(hr + 1, 1)
    return ws


wb = Workbook()
wb.remove(wb.active)
'''

FOOTER = '''
# mandatory fields highlighted yellow
ws = wb["0_Parameters"]
for row in ws.iter_rows(min_row=4):
    if str(row[0].value).strip() in ("brandklasse", "signaalfamilie"):
        row[1].fill = PatternFill("solid", start_color="FFFF00")

wb.save("config/kabelconfig.xlsx")
print("config/kabelconfig.xlsx written")
'''


def main():
    wb = load_workbook("config/kabelconfig.xlsx")
    out = [HEADER]
    for name in wb.sheetnames:
        ws = wb[name]
        data = [[c for c in r] for r in ws.iter_rows(values_only=True)]
        note, i = None, 0
        if (data and data[0] and data[0][0]
                and sum(1 for c in data[0] if c not in (None, "")) == 1
                and len(str(data[0][0])) > 40):
            note = str(data[0][0]); i = 1
            while i < len(data) and all(c in (None, "") for c in data[i]):
                i += 1
        headers = [c for c in data[i] if c not in (None, "")]
        ncol = len(data[i])
        body = [tuple(("" if c is None else c) for c in r[:ncol])
                for r in data[i + 1:] if r and any(c not in (None, "") for c in r)]
        widths = [max(10, min(60, max([len(str(headers[j] or ""))]
                  + [len(str(b[j])) for b in body if j < len(b)]) + 2))
                  for j in range(len(headers))]
        out.append(f"sheet(wb, {name!r},\n    {note!r},\n    {headers!r},\n    [")
        out += [f"        {b!r}," for b in body]
        out.append(f"    ],\n    {widths!r})\n")
    out.append(FOOTER)
    open("build_config.py", "w").write("\n".join(out))
    print("build_config.py regenerated from config/kabelconfig.xlsx")


if __name__ == "__main__":
    main()

"""Bootstrap a fresh config/kabelconfig.xlsx.

AUTO-GENERATED from the live workbook — do not hand-edit. Regenerate with:
    python3 dump_config.py > build_config.py

The workbook is the source of truth and the estimators edit it directly. This
script exists only to recreate it from scratch (new install, or recovering a
corrupted file). Running it OVERWRITES config/kabelconfig.xlsx — take a copy of
any project parameters first.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

HDR = Font(bold=True, color="FFFFFF", name="Arial", size=10)
FILL = PatternFill("solid", start_color="1F4E78")
BODY = Font(name="Arial", size=10)
NOTE = Font(name="Arial", size=9, italic=True, color="666666")


def sheet(wb, name, note, headers, rows, widths):
    ws = wb.create_sheet(name)
    if note:
        ws.append([note]); ws.cell(1, 1).font = NOTE
        ws.append([])
    ws.append(headers)
    hr = ws.max_row
    for c in range(1, len(headers) + 1):
        cell = ws.cell(hr, c)
        cell.font, cell.fill = HDR, FILL
        cell.alignment = Alignment(vertical="center")
    for r in rows:
        ws.append(list(r))
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    for row in ws.iter_rows(min_row=hr + 1):
        for c in row:
            c.font = BODY
    ws.freeze_panes = ws.cell(hr + 1, 1)
    return ws


wb = Workbook()
wb.remove(wb.active)

sheet(wb, '0_Parameters',
    "Step 0 — fill in per project BEFORE generation. Yellow fields are mandatory. Values in 'opties' (geen/erco/derden) are literal keys read by the engine — keep them as-is.",
    ['parameter', 'waarde', 'opties', 'toelichting'],
    [
        ('project_nr', 'PR 20267283-2600206', '', 'Erco PR number'),
        ('project_naam', 'Duitslandlaan Zoetermeer', '', ''),
        ('rk_naam', 'RK071', 'AUTO or fixed (e.g. RK071)', 'AUTO = derived from the process-code prefixes'),
        ('plaats_rk', 'Technische ruimte begane grond', '', 'human input, NOT present in the I/O list'),
        ('brandklasse', 'B2CA', 'CCA / B2CA', 'MANDATORY — not derivable from the I/O list (Duitslandlaan + 112-meldkamer lesson)'),
        ('signaalfamilie', 'JOBA', 'DRAK / JOBA', 'CCA -> usually DRAK, B2CA -> usually JOBA; confirm per project'),
        ('lengteklasse', '<=50m', '<=50m / 50-100m', ">50m -> JOBA variant + 'max ..m' note"),
        ('naregelsysteem', 'geen', 'geen / BRN15 / Priva / Touchpoint', 'determines the bus cable + room-control item set'),
        ('tracing_scope', 'erco', 'erco / derden', 'Duitslandlaan precedent: Erco feeds from the RK, despite the input remark'),
        ('locatie_veld_standaard', 'in TR', '', "default 'bekabeling naar' for field cables"),
        ('regelkast_bestaand', 'nee', 'ja / nee', "renovation: existing panel -> 'Bestaande regelkast geen aanpassingen', no RK feed cable"),
        ('wp_aansluiten_erco', 'nee', 'ja / nee', "ja: WP feed = 'Kabel levering derde totaan WP, aansluiten WPzijde Erco' (counts in derden total)"),
        ('brandmelding_standaard', 'ja', 'ja / nee', 'nee for existing installations without a new fire-alarm cable (Boerhaave)'),
    ],
    [24, 32, 35, 60])

sheet(wb, '1_Synoniemen',
    "Step 2 — classification dictionary. The longest / highest-priority match wins. Add every manual correction here (the system gets smarter with each project). The 'functietype' codes are keys shared with tab 2_Kabelkeuze — keep them identical.",
    ['patroon (lowercase, substring)', 'functietype', 'prioriteit', 'source/example'],
    [
        ('modbus ip', 'BUS_MODBUS', 90, 'Duitslandlaan AIR-SEP: Modbus-IP field device -> BMS cable'),
        ('modbus', 'BUS_MODBUS', 50, ''),
        ('bacnet ms/tp', 'BUS_MODBUS', 60, ''),
        ('bacnet', 'BUS_IP_PANEEL', 55, '112: BACnet/IP -> COMM CAT6, per device, no daisy-chain'),
        ('m-bus', 'BUS_MBUS', 60, ''),
        ('buskoppeling', 'BUS_MODBUS', 40, ''),
        ('mod-bus', 'BUS_MODBUS', 60, ''),
        ('communicatie interface', 'BUS_INTERFACE', 70, 'CIM200 print: belongs to the pump, generates no cable of its own'),
        ('touchpanel', 'BUS_IP_PANEEL', 70, ''),
        ('kwh-meter', 'ENERGIEMETER', 60, ''),
        ('energiemeter', 'ENERGIEMETER', 60, 'RH33: bus + separate 24V feed cable'),
        ('koudemeter', 'ENERGIEMETER', 60, ''),
        ('warmtemeter', 'ENERGIEMETER', 60, ''),
        ('flowmeting', 'METING_ACTIEF', 70, 'prio raised above energiemeter tie'),
        ('flowmeter', 'METING_ACTIEF', 70, ''),
        ('optie c1', 'METER_OPTIE_VOELER', 90, "PT100 of an E+H meter -> 'Via aansluitsnoer', no cable"),
        ('pt100', 'METER_OPTIE_VOELER', 40, 'only when tied to an energy meter; otherwise METING_PASSIEF'),
        ('drukverschil', 'METING_ACTIEF', 60, '22WDP -> 5X1/2502Q'),
        ('systeemdruk', 'METING_ACTIEF', 60, '22WP-119'),
        ('druk', 'METING_ACTIEF', 30, ''),
        ('temperatuur', 'METING_PASSIEF', 65, 'prio raised above warmtepomp(40) — Boerhaave over-match fix'),
        ('buitenvoeler', 'METING_ACTIEF', 60, 'only if present in the input (lesson: do not add it yourself)'),
        ('blokkeerafsluiter', 'KLEP_STURING_MELDING', 70, 'GR24A-SR: control + feedback -> 7X1'),
        ('regelafsluiter open/dicht', 'KLEP_STURING_MELDING', 80, ''),
        ('regelafsluiter', 'REGELAFSLUITER_0_10V', 50, '0-10V -> 5X1 / RAK 4X0,8'),
        ('klep', 'KLEP_OD', 30, ''),
        ('brandklep', 'BRANDKLEP', 70, '112: one 2-core cable PER damper, do not bundle'),
        ('vrijgave', 'VRIJGAVE', 50, ''),
        ('sturing 0-10', 'STURING_0_10V', 60, ''),
        ('sturing', 'STURING_0_10V', 30, 'E-boiler: AO on its own 2-core cable'),
        ('bedrijf/storing', 'BEDRIJF_STORING', 60, ''),
        ('bedrijfsmeldingen', 'MELDINGEN_GROOT', 70, 'WP: all meldingen on one 12X1'),
        ('diverse meldingen', 'MELDINGEN_GROOT', 70, ''),
        ('storing', 'MELDING', 30, 'single contact -> 2X1/2X0,8'),
        ('melding', 'MELDING', 25, ''),
        ('tracing', 'TRACING', 70, 'feed from the RK (ws=1) + 2-core feedback'),
        ('warmtepomp', 'WARMTEPOMP', 40, 'bus + meldingen bundle; feed = Werkzaamheden derden'),
        ('elektrische ketel', 'E_KETEL', 70, 'vrijgave/storing 5X1 + separate sturing 2X1; feed by third parties'),
        ('levering derden', 'DERDEN', 80, ''),
        ('levering derde', 'DERDEN', 80, ''),
        ('vanuit bmc', 'DERDEN', 80, ''),
        ('uit e-installatie', 'DERDEN', 80, ''),
        ('vanuit e-verdeler', 'DERDEN', 80, ''),
        ('vanuit hvk', 'DERDEN', 80, ''),
        ('voeding door derden', 'DERDEN', 85, ''),
        ('regelkast', 'REGELKAST', 90, 'own feed = Kabel levering derde totaan RK'),
        ('n.t.b.', 'NTB', 85, 'unknown scope -> omit + flag'),
        ('brandmelding', 'BRANDMELDING', 80, 'always present in Onderstation algemeen'),
        ('circulatiepomp', 'POMP', 50, ''),
        ('circulaite pomp', 'POMP', 50, 'typo variant seen in the input'),
        ('circ.pomp', 'POMP', 50, ''),
        ('transportpomp', 'POMP', 50, '400V without N -> 4G2,5 (NOT 5G)'),
        ('pomp', 'POMP', 25, ''),
        ('ontgasser', 'APPARAAT_230V', 50, 'Korex AIR-SEP: feed + fault + bus'),
        ('ventilator', 'EC_VENTILATOR', 40, 'run/fault/control -> 3X2 AFG'),
        ('elektrameter', 'EVERDELER_METER', 95, 'Boerhaave: meter in the E-verdeler -> bus row in Onderstation algemeen'),
        ('smoorafsluiter', 'SMOORAFSLUITER', 80, 'Boerhaave: SAL81 + hulprelais -> JOBA STUURSTR 7X1'),
        ('leidingverwarming', 'TRACING', 75, 'synonym for tracing'),
        ('intredetemp', 'METING_PASSIEF', 65, ''),
        ('uittredetemp', 'METING_PASSIEF', 65, ''),
        ('aanvoertemperatuur', 'METING_PASSIEF', 66, ''),
        ('retourtemperatuur', 'METING_PASSIEF', 66, ''),
    ],
    [32, 22, 12, 60])

sheet(wb, '2_Kabelkeuze',
    'Step 3 — cable choice per function type per family. One family switch = one parameter. Strings are verbatim as they appear on the cable list.',
    ['functietype', 'kabel_CCA_DRAK', 'kabel_B2CA_JOBA', 'opmerking'],
    [
        ('METING_PASSIEF', 'DRAK SIGK CCA 1X2X0,8 2501 MT', 'JOBA ST.STR B2CA HCHOZ 2X1 MT', 'passive temperature sensor'),
        ('MELDING', 'DRAK CCA GY 2X0,8 MT', 'JOBA ST.STR B2CA HCHOZ 2X1 MT', 'single volt-free contact'),
        ('STURING_0_10V', 'RAK SIGN KAB CCA 4X0,8 HA500', 'JOBA ST.STR B2CA HCHOZ 2X1 MT', 'standalone AO (E-boiler sturing = 2X1!)'),
        ('TRACING_TERUGMELDING', 'DRAK CCA GY 2X0,8 MT', 'JOBA ST.STR B2CA HCHOZ 2X1 MT', ''),
        ('METING_ACTIEF', 'DRAK SIGK CCA 1X4X0,8 2502Q MT', 'JOBA ST.STR B2CA HCHJZ 5X1 MT', 'powered sensor, feed+signal on 1 cable up to ~30VA'),
        ('METER_VOEDING_24V', 'JOBA STUURSTR HHJZ 3X1 MT', 'JOBA ST.STR B2CA HCHJZ 5X1 MT', 'separate feed cable energy meter — CCA per Boerhaave (3X1), B2CA per Duitslandlaan (5X1)'),
        ('REGELAFSLUITER_0_10V', 'RAK SIGN KAB CCA 4X0,8 HA500', 'JOBA ST.STR B2CA HCHJZ 5X1 MT', ''),
        ('VRIJGAVE_STORING', 'DRAK SIGK CCA 1X4X0,8 2502Q MT', 'JOBA ST.STR B2CA HCHJZ 5X1 MT', 'E-boiler enable/fault'),
        ('BEDRIJF_STORING', 'RAK SIGN KAB CCA 4X0,8 HA500', 'JOBA ST.STR B2CA HCHJZ 5X1 MT', ''),
        ('KLEP_STURING_MELDING', 'DRAK CCA GY 6X0,8 MT', 'JOBA STSTR B2CA HCHJZ 7X1 MT', 'isolation valve control + feedback'),
        ('KLEP_OD', 'DRAK CCA GY 8X0,8 MT', 'JOBA ST.STR B2CA HCHJZ 7X1 MT', 'open/close + feedback(s)'),
        ('BRANDKLEP', 'DRAK CCA GY 2X0,8 MT', 'JOBA ST.STR B2CA HCHOZ 2X1 MT', 'one cable PER damper (112)'),
        ('MELDINGEN_GROOT', 'JOBA HCH-JZ 12X1 B2CA MT', 'JOBA HCH-JZ 12X1 B2CA MT', "WP meldingen, fan 'diverse functions'"),
        ('EC_VENTILATOR', 'DRAK SIGK AFG CCA 3X2X0,8 MT', 'JOBA ST.STR B2CA HCHJZ 7X1 MT', 'run/fault/control'),
        ('SMOORAFSLUITER', 'JOBA STUURSTR HHJZ 7X1 MT', 'JOBA STSTR B2CA HCHJZ 7X1 MT', 'throttle valve open/close + feedback (Boerhaave CCA validated)'),
    ],
    [22, 32, 31, 60])

sheet(wb, '3_Voedingen',
    "Feed rules. 400V frequency-controlled pumps (TPE3) = without N -> 4G2,5. Only cables that arrive at the RK count toward 'Totaal voedingen derden aansluiten'. Class keys (230V_1F etc.) and the house strings are read literally — keep them as-is.",
    ['klasse', 'kabel_CCA', 'kabel_B2CA', 'ws', 'bekabeling_naar_sjabloon', 'ws_CCA', 'ws_B2CA', 'sjabloon_CCA', 'sjabloon_B2CA'],
    [
        ('230V_1F', 'DRAK HULT CCA 3G2,5 HA500', 'DRAK HULT B2CA 3G2,5 MT', 1, '{kw}kW/{a}A/230V', '', '', '', ''),
        ('400V_3F_zonder_N', 'DRAK HULT CCA 4X2,5 MT', 'DRAK HULT B2CA 4G2,5 MT', 1, '{kw}kW/{a}A/400V', '', '', '', ''),
        ('400V_3F_met_N', 'DRAK HULT CCA 5G2,5 MT', 'DRAK HULT B2CA 5G2,5 MT', 1, '{kw}kW/{a}A/400V', '', '', '', ''),
        ('TRACING', 'DRAK HULT CCA 5G2,5 MT', 'DRAK HULT B2CA 3G2,5 MT', '', '', '', 1, 'voeding doorlussen', ''),
        ('DERDEN_RK', 'Kabel levering derde totaan RK, aansluiten kastzijde Erco', 'Kabel levering derde totaan RK, aansluiten kastzijde Erco', '', '', '', '', '', ''),
        ('DERDEN_TOESTEL', 'Werkzaamheden derden', 'Werkzaamheden derden', '', '', '', '', '', ''),
    ],
    [18, 59, 59, 10, 26, 10, 10, 20, 15])

sheet(wb, '4_Bus',
    "Bus cables per protocol. The 'doorlus_gedrag' values (doorlussen/geen) are literal keys.",
    ['bustype', 'kabel', 'doorlus_gedrag', 'opmerking'],
    [
        ('MODBUS_RTU', 'BMS Cable 2x2x24AWG - R1319 - B2ca s1,d0,a1 Violet HA500', 'doorlussen', "1st device: location; 2nd+: 'doorlussen' (lowercase)"),
        ('MODBUS_IP_VELD', 'BMS Cable 2x2x24AWG - R1319 - B2ca s1,d0,a1 Violet HA500', 'geen', 'Modbus-IP field device = BMS cable, NO UTP (Duitslandlaan Korex)'),
        ('BACNET_IP', 'COMM U/UTP CAT6 CS34ZB HA305', 'geen', 'one cable per device, no daisy-chain (112)'),
        ('TOUCHPANEL_IP', 'GGM C6U4PFROHT3 GGM U/UTP6 LSZH 305M', 'geen', ''),
        ('RK_ONDERLING', 'COMM U/UTP CAT6 CS34ZC HA305', 'geen', 'or CS34ZB, depending on the room-control system'),
        ('MBUS_DERDEN', 'Kabel levering derde, enkelzijdig aansluiten op regelkast Erco', 'geen', ''),
    ],
    [16, 60, 16, 60])

sheet(wb, '5_Vaste_teksten',
    "Literal house texts. These land verbatim in the cable list — do not translate the 'tekst' column.",
    ['sleutel', 'tekst'],
    [
        ('AANSLUITSNOER', 'Via aansluitsnoer van 2 meter op meter'),
        ('BRANDMELDING', 'Kabel levering derde totaan RK, aansluiten kastzijde Erco'),
        ('NAREGEL_UNIT_DERDE', 'Kabel levering derde, aanbrengen werkzaamheden Erco'),
        ('TOT_DERDEN', 'Totaal voedingen derden aansluiten'),
        ('TOT_WS', 'Totaal aantal werkschakelaars van maximaal 63A'),
        ('ONDERSTATION', 'Onderstation algemeen'),
        ('VOEDINGEN', 'Voedingen'),
        ('BESTAAND_RK', 'Bestaande regelkast geen aanpassingen'),
        ('VOEDING_REEDS', 'Voeding reeds aangesloten'),
        ('DERDEN_WP', 'Kabel levering derde totaan WP, aansluiten WPzijde Erco'),
    ],
    [20, 59])

sheet(wb, '6_Locatiekoppen',
    "Group headers in the input that are LOCATION banners, not equipment sections. Matching rows get 'locatie' in the bekabeling-naar column and are sectioned by their equipment instead. Evidence: Fonkel PR20267276 -- input 'Installaties op het dak' -> manual section 'Warmtepomp' with 'op dak'; 'Installaties buiten bij buffervat' -> manual section 'Buffervat CV' with 'Buiten TR'.",
    ['patroon (substring, lowercase)', 'locatie', 'sectie_fallback'],
    [
        ('installaties op het dak', 'op dak', 'Warmtepomp'),
        ('op het dak', 'op dak', 'Warmtepomp'),
        ('installaties buiten bij buffervat', 'Buiten TR', 'Buffervat CV'),
        ('buiten bij', 'Buiten TR', ''),
        ('in technische ruimte', 'in TR', ''),
    ],
    [35, 11, 17])


# mandatory fields highlighted yellow
ws = wb["0_Parameters"]
for row in ws.iter_rows(min_row=4):
    if str(row[0].value).strip() in ("brandklasse", "signaalfamilie"):
        row[1].fill = PatternFill("solid", start_color="FFFF00")

wb.save("config/kabelconfig.xlsx")
print("config/kabelconfig.xlsx written")

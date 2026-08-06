"""Regression 3: Fonkel Breda (PR 20267276-2600214) — B2CA/JOBA, Kuijpers layout.
Validated against the manual cable list of 10-7-2026 (Rick van Deurzen).

Input: Functielijst Fonkel Breda (PREC000155, 7-7-2026), transcribed to
projects/fonkel/normalized.csv.

Establishes the LOCATION-BANNER rule: an input group header may be a location
("Installaties op het dak", "Installaties buiten bij buffervat") rather than an
equipment section. Such a group sets 'bekabeling naar' and the section falls back
to the equipment name — exactly how the estimator reads it. Contrast Boerhaave,
whose groups ARE equipment names and are used verbatim.

DOCUMENTED DEVIATIONS (accepted, not bugs):
- `Totaal ws`: engine computes 8 from the ws column; the manual shows 0. Same
  unexplained contradiction as Duitslandlaan (10 vs 0). ASK — see CHANGELOG.
- Tracing sub-panel: the manual has a separate "Tracing" page (own Voedingen,
  derden=0, 2 cables, no RK code). Not yet modelled — batched.
- Row labels differ cosmetically ("Temperatuur hoog buffer LT-CV" vs
  "Buffervattemperatuuropnemer LT-CV hoog"); cable and location identical.
- Blokkeerafsluiter Serie-/Parallelbedrijf rows (081RA_11/12 duplicates) are
  omitted from the fixture: the manual lists 4 regelafsluiters total, matching
  the transport valves only.
  - Tracing feed: the manual's separate "Tracing" page carries 1 feed + 2 cables.
  Evaluation is on the set of cables, not their page, so the tracing voeding is
  expected in this list. Counts raised 2->3 (op dak) and 4->5 (3G2,5). Previously
  masked: 'optioneel vanuit regelkast' misclassified the row as a panel row,
  which suppressed the feed via the REGELKAST early-return.
"""
import subprocess
import sys
from openpyxl import load_workbook

subprocess.run([sys.executable, "-m", "io2cable.pipeline",
                "--config", "config/kabelconfig_fonkel.xlsx",
                "--input", "projects/fonkel/normalized.csv",
                "--out", "projects/fonkel/out_test", "--rk", "RK071"],
               check=True, capture_output=True)

ws = load_workbook("projects/fonkel/out_test/cable_list_RK071.xlsx").active
rows = [tuple("" if c is None else str(c) for c in r) for r in ws.iter_rows(values_only=True)]
text = "\n".join(" | ".join(r) for r in rows)

sections = [r[1] for r in rows if r[1] and not r[0] and "Totaal" not in r[1] and r[1] != "onderdeel"]

def has(needle, n=None):
    cnt = text.count(needle)
    ok = (cnt >= 1) if n is None else (cnt == n)
    print(("PASS" if ok else "FAIL"), needle[:64], f"(x{cnt}{'' if n is None else f', expected {n}'})")
    return ok

checks = []


# --- Voedingen + totals ----------------------------------------------------
checks.append(has("Kabel levering derde totaan RK, aansluiten kastzijde Erco", 2))  # RK + brandmelding
checks.append(has("Werkzaamheden derden", 2))                                       # WP + E-ketel
checks.append(has("Totaal voedingen derden aansluiten | 1"))                        # dedup fix

# --- cable strings (B2CA/JOBA) --------------------------------------------
checks.append(has("JOBA ST.STR B2CA HCHOZ 2X1 MT", 15))   # 7 buffer + 2 centrale + 4 TSA + korex storing + ketel sturing
checks.append(has("JOBA ST.STR B2CA HCHJZ 5X1 MT", 13))   # druk/flow/pdt/EM-voeding/ketel
checks.append(has("JOBA STSTR B2CA HCHJZ 7X1 MT", 4))     # blokkeerafsluiters
checks.append(has("JOBA HCH-JZ 12X1 B2CA MT", 1))         # WP meldingen — ONE, not two
checks.append(has("DRAK HULT B2CA 4G2,5 MT", 4))          # 400V transport pumps
checks.append(has("DRAK HULT B2CA 3G2,5 MT", 5))          # 230V circ/deelstroom/korex
checks.append(has("Via aansluitsnoer van 2 meter op meter", 4))
checks.append(has("BMS Cable 2x2x24AWG - R1319 - B2ca s1,d0,a1 Violet HA500", 11))

# --- device-level rules ----------------------------------------------------
checks.append(has("Elektrische ketel vrijgave/storing"))
checks.append(has("Elektrische ketel sturing"))
checks.append(has("Brandmelding"))

print(f"\n{sum(checks)}/{len(checks)} checks passed")
sys.exit(0 if all(checks) else 1)

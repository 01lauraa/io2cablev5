"""Regression test 2: Boerhaave Leiden RK03 (PR 20267275.1) — CCA family,
renovation project, multi-WP. Validated against the manual cable list of 15-7-2026.

DOCUMENTED DEVIATIONS from the manual PDF (accepted, pending estimator answers):
- WP1 has a MODbus row here (+1 BMS cable); the manual list omits it (existing
  cabling on the renovated panel — human review decision).
- Chain labels: engine writes 'doorlussen'; manual list uses 'Uit WP 1'/'Uit WP 2'
  and 'maximaal 40 meter' (physical routing/length = human input, §6 ambiguity).
- Row labels differ slightly (e.g. 'Energiemeter WP 24V voeding' vs 'Energiemeter
  voeding') — content and cable identical.
Run: python3 test_regression_boerhaave.py"""
import subprocess
import sys
from openpyxl import load_workbook

subprocess.run([sys.executable, "-m", "io2cable.pipeline",
                "--config", "config/kabelconfig_boerhaave.xlsx",
                "--input", "projects/boerhaave/normalized.csv",
                "--out", "projects/boerhaave/out_test", "--rk", "RK03"],
               check=True, capture_output=True)

ws = load_workbook("projects/boerhaave/out_test/cable_list_RK03.xlsx").active
rows = [tuple("" if c is None else str(c) for c in r) for r in ws.iter_rows(values_only=True)]
text = "\n".join(" | ".join(r) for r in rows)

def has(needle, n=None):
    cnt = text.count(needle)
    ok = (cnt >= 1) if n is None else (cnt == n)
    print(("PASS" if ok else "FAIL"), needle[:70], f"(x{cnt}{'' if n is None else f', expected {n}'})")
    return ok

checks = [
    # renovation Voedingen states
    has("Bestaande regelkast geen aanpassingen", 1),
    has("Voeding reeds aangesloten", 1),                                   # WP1
    has("Kabel levering derde totaan WP, aansluiten WPzijde Erco", 2),     # WP2+WP3
    # totals: derden counts Erco terminations (2), ws none in this project
    has("Totaal voedingen derden aansluiten | 2"),
    has("Totaal aantal werkschakelaars van maximaal 63A | 0"),
    # mixed-family house standard within one CCA project
    has("JOBA STUURSTR HHJZ 7X1 MT", 12),         # 4 smoorafsluiters x 3 WP
    has("JOBA STUURSTR HHJZ 3X1 MT", 3),          # energiemeter voeding per WP
    has("DRAK SIGK CCA 1X2X0,8 2501 MT", 7),      # buffer temps
    has("DRAK HULT CCA 5G2,5 MT", 3),             # tracing per WP
    # bus: 3 WP + 3 energiemeter + 3 E-verdeler = 9 (manual: 8; WP1 deviation documented)
    has("BMS Cable 2x2x24AWG - R1319 - B2ca s1,d0,a1 Violet HA500", 9),
    # E-verdeler meters land in Onderstation algemeen; NO brandmelding (param nee)
    has("Onderstation algemeen"),
    has("Elektrameter warmtepomp MODbus koppeling", 3),
]
absent = "Brandmelding" not in text
print(("PASS" if absent else "FAIL"), "no Brandmelding row (brandmelding_standaard=nee)")
checks.append(absent)
absent2 = "Werkzaamheden derden" not in text
print(("PASS" if absent2 else "FAIL"), "no 'Werkzaamheden derden' (this project uses WPzijde Erco)")
checks.append(absent2)

print(f"\n{sum(checks)}/{len(checks)} checks passed")
sys.exit(0 if all(checks) else 1)

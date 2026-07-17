"""Regression: <Project> (<PR>) — <family>, <layout>.
Validated against the manual cable list of <date>.

DOCUMENTED DEVIATIONS (accepted, not bugs):
- <e.g. chain labels: engine writes 'doorlussen', manual uses 'Uit WP 1'>
- <e.g. 'maximaal 40 meter' — length is human input, §6 ambiguity>
"""
import subprocess, sys, os
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
NAME = os.path.basename(HERE)

subprocess.run([sys.executable, "-m", "io2cable.pipeline",
                "--config", f"config/kabelconfig_{NAME}.xlsx",
                "--input",  f"projects/{NAME}/normalized.csv",
                "--out",    f"projects/{NAME}/out_test", "--rk", "RK01"],
               check=True, capture_output=True)

ws = load_workbook(f"projects/{NAME}/out_test/cable_list_RK01.xlsx").active
rows = [tuple("" if c is None else str(c) for c in r) for r in ws.iter_rows(values_only=True)]
text = "\n".join(" | ".join(r) for r in rows)

def has(needle, n=None):
    cnt = text.count(needle)
    ok = (cnt >= 1) if n is None else (cnt == n)
    print(("PASS" if ok else "FAIL"), needle[:66], f"(x{cnt}{'' if n is None else f', expected {n}'})")
    return ok

checks = [
    # totals
    has("Totaal voedingen derden aansluiten | ?"),
    has("Totaal aantal werkschakelaars van maximaal 63A | ?"),
    # cable strings with exact counts
    # has("<CABLE STRING>", <n>),
]

print(f"\n{sum(checks)}/{len(checks)} checks passed")
sys.exit(0 if all(checks) else 1)

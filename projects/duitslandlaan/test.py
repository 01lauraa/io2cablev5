"""Regression 1: Duitslandlaan Zoetermeer (PR 20267283) — B2CA/JOBA, Kuijpers layout.
Validated against the manual cable list (Rick van Deurzen).

Established brandklasse as a mandatory, non-derivable parameter: defaulting it to
CCA produced ~35 wrong rows from one wrong cell.

DOCUMENTED DEVIATIONS (accepted, not bugs):
- `Totaal ws`: engine computes 10 from the ws column; the manual shows 0. Same
  unexplained contradiction as Fonkel (8 vs 0). ASK — see docs/CHANGELOG.md.
- Daisy-chain order assumes input order; physical routing is human input (§6).
"""
import subprocess
import sys
from openpyxl import load_workbook

subprocess.run([sys.executable, "-m", "io2cable.pipeline",
                "--config", "config/kabelconfig_duitslandlaan.xlsx",
                "--input", "projects/duitslandlaan/normalized.csv",
                "--out", "projects/duitslandlaan/out_test", "--rk", "RK071"],
               check=True, capture_output=True)

ws = load_workbook("projects/duitslandlaan/out_test/cable_list_RK071.xlsx").active
rows = [tuple("" if c is None else str(c) for c in r) for r in ws.iter_rows(values_only=True)]
text = "\n".join(" | ".join(r) for r in rows)
cables = [r for r in rows if r[0].isdigit() and r[4]]

def has(needle, n=None):
    cnt = text.count(needle)
    ok = (cnt >= 1) if n is None else (cnt == n)
    print(("PASS" if ok else "FAIL"), needle, f"(x{cnt}{'' if n is None else f', expected {n}'})")
    return ok

checks = [
    # totals per rules v2 (derden = only RK-arriving; ws from column)
    has("Totaal voedingen derden aansluiten | 1"),
    has("Totaal aantal werkschakelaars van maximaal 63A | 10"),
    # voedingen strings
    has("Kabel levering derde totaan RK, aansluiten kastzijde Erco", 2),  # RK + brandmelding
    has("Werkzaamheden derden", 2),                                        # WP + E-ketel
    # cable-family mapping (B2CA/JOBA project)
    has("JOBA ST.STR B2CA HCHOZ 2X1 MT", 17),   # 14 temps + korex fault + tracing feedback + boiler control
    has("JOBA ST.STR B2CA HCHJZ 5X1 MT", 14),   # 2 PT + 2 FT + 2 EM feed + 6 PDT + RA01 + boiler enable/fault
    has("JOBA STSTR B2CA HCHJZ 7X1 MT", 4),     # blokkeerafsluiters
    has("JOBA HCH-JZ 12X1 B2CA MT", 1),         # WP meldingen
    has("DRAK HULT B2CA 4G2,5 MT", 4),          # 400V transport pumps (NOT 5G!)
    has("DRAK HULT B2CA 3G2,5 MT", 6),          # tracing + korex + 2 partial-flow + 2 circ
    has("BMS Cable 2x2x24AWG - R1319 - B2ca s1,d0,a1 Violet HA500", 12),  # WP + 2 EM + korex + 8 pumps (kWh meter omitted)
    has("Via aansluitsnoer van 2 meter op meter", 4),
    has("doorlussen", 10),
    # structure
    has("Elektrische ketel vrijgave/storing"),
    has("Elektrische ketel sturing"),
    has("Brandmelding"),
]
# kWh meter (n.t.b.) must be OMITTED
absent = "kWh" not in text
print(("PASS" if absent else "FAIL"), "kWh meter (n.t.b.) omitted")
checks.append(absent)

n_cables = len(cables)
ok_n = n_cables == 66  # 3 feeds + 62 device cables + fire alarm, matching the real list
print(("PASS" if ok_n else "FAIL"), f"66 cable rows (actual {n_cables})")
checks.append(ok_n)

print(f"\n{sum(checks)}/{len(checks)} checks passed")
sys.exit(0 if all(checks) else 1)

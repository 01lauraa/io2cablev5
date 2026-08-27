"""Pipeline runner: Step 1 → review file → Step 2 → Step 3 → kabellijst xlsx + flags.
Usage:
    python -m io2cable.pipeline --config config/kabelconfig.xlsx \
        --input fixtures/duitslandlaan_normalized.csv --out out/
Input may be a function list (functielijst) .xlsx (Layer A parse) or a normalized .csv
(Layer B / reviewed).
"""
import argparse
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from .config import load_config, load_header_map
from . import ingest, classify, rules

BOLD = Font(bold=True, name="Arial", size=10)
BLUE = Font(bold=True, italic=True, color="1F4E78", name="Arial", size=10)


def write_cable_list(result, cfg, path, rk_naam):
    wb = Workbook()
    ws = wb.active
    ws.append(["nr", "onderdeel", "ws", "proces code", "kabel soort en doorsnede"])
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        c.fill = PatternFill("solid", start_color="1F4E78")
    ws.append([])
    ws.append(["", cfg.texts["VOEDINGEN"]]); ws.cell(ws.max_row, 2).font = BLUE
    for i, r in enumerate(result["voedingen"], 1):
        ws.append([i, r.onderdeel, r.ws, r.procescode, r.kabel])
    ws.append([])
    ws.append(["", cfg.texts["TOT_DERDEN"], result["tot_derden"]])
    ws.append(["", cfg.texts["TOT_WS"], result["tot_ws"]])
    for rr in (ws.max_row - 1, ws.max_row):
        ws.cell(rr, 2).font = BOLD
        ws.cell(rr, 3).font = Font(bold=True, color="0000FF", name="Arial", size=10)
    ws.append([])
    nr = 0
    current_section = None
    for r in result["devices"]:
        if r.section != current_section:
            if current_section is not None:
                ws.append([])
            ws.append(["", r.section]); ws.cell(ws.max_row, 2).font = BLUE
            current_section = r.section
        nr += 1
        flag = " [?]" if r.flags else ""
        ws.append([nr, r.onderdeel + flag, r.ws, r.procescode, r.kabel])
    ws.append([])
    ws.append(["", cfg.texts["ONDERSTATION"]]); ws.cell(ws.max_row, 2).font = BLUE
    for r in result["onderstation"]:
        nr += 1
        ws.append([nr, r.onderdeel, r.ws, r.procescode, r.kabel])
    for col, w in zip("ABCDE", (5, 46, 5, 12, 58)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    wb.save(path)


def main(argv=None):
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", required=True)
    ap.add_argument("--input", required=True, help=".xlsx function list or normalized .csv")
    ap.add_argument("--header-map", default=None,
                    help="config/header_map_<client>.xlsx for a non-standard column layout")
    ap.add_argument("--sheet", default=None,
                    help="worksheet name; default = the workbook's active sheet. "
                         "Multi-sheet functielijsten (RK01/RK02/N) are run one sheet at a time.")
    ap.add_argument("--list-sheets", action="store_true",
                    help="print the worksheet names of --input and exit")
    ap.add_argument("--out", default="out")
    ap.add_argument("--rk", default=None)
    args = ap.parse_args(argv)

    if args.list_sheets:
        from openpyxl import load_workbook as _lw
        for s in _lw(args.input, read_only=True).sheetnames:
            print(s)
        return

    cfg = load_config(args.config)
    missing = [p for p in ("brandklasse",) if not str(cfg.parameters.get(p, "")).strip()]
    if missing:
        raise SystemExit(f"STOP: mandatory parameter(s) empty in 0_Parameters: {missing}. "
                         "Fire class (brandklasse) cannot be derived from the I/O list (Duitslandlaan lesson).")

    os.makedirs(args.out, exist_ok=True)
    if args.input.lower().endswith(".csv"):
        norm = ingest.load_normalized(args.input)
    else:
        hmap = load_header_map(args.header_map) if args.header_map else None
        norm = ingest.parse_excel(args.input, rk=args.rk or "RK?", header_map=hmap,
                                  sheet=args.sheet)
    ingest.write_review(norm, os.path.join(args.out, "normalized_review.xlsx"))

    classified = classify.classify(norm, cfg)
    results, flags = rules.run_per_rk(classified, cfg)

    for rk_key, result in results.items():
        rk = args.rk or str(cfg.parameters.get("rk_naam") or "") or rk_key
        if len(results) > 1:
            rk = rk_key  # multiple panels: each keeps its own name from the input
        if rk == "AUTO":
            pref = sorted({r.sort_key[0] for r in result["devices"]})
            rk = "RK" + "-".join(f"{q:03d}" for q in pref[:1]) if pref else "RK"
        write_cable_list(result, cfg, os.path.join(args.out, f"cable_list_{rk}.xlsx"), rk)

    with open(os.path.join(args.out, "flags.txt"), "w", encoding="utf-8") as f:
        seen = set()
        for fl in flags:
            if fl not in seen:
                f.write(fl + "\n")
                seen.add(fl)
    summary = "; ".join(
        f"{k}: {len(v['voedingen'])} feeds + {len(v['devices'])} cables, "
        f"derden={v['tot_derden']}, ws={v['tot_ws']}" for k, v in results.items())
    print(f"OK: {len(norm)} input rows -> {summary}; {len(set(flags))} flags -> {args.out}/")


if __name__ == "__main__":
    main()

"""Loads kabelconfig.xlsx into runtime structures. The workbook is the source of
truth the estimators maintain; nothing cable-related is hardcoded in Python."""
from dataclasses import dataclass, field
from openpyxl import load_workbook


def _rows(ws):
    """Yield dict rows using the header row (first bold-ish row with >1 filled cell after notes)."""
    data = list(ws.iter_rows(values_only=True))
    hdr_i = next(i for i, r in enumerate(data)
                 if r and sum(1 for c in r if c not in (None, "")) > 1)
    headers = [str(c).strip() if c else "" for c in data[hdr_i]]
    for r in data[hdr_i + 1:]:
        if not r or all(c in (None, "") for c in r):
            continue
        yield {h: (c if c is not None else "") for h, c in zip(headers, r)}


@dataclass
class Config:
    parameters: dict = field(default_factory=dict)
    synonyms: list = field(default_factory=list)      # (pattern, functietype, prio)
    cable_by_type: dict = field(default_factory=dict) # functietype -> {CCA, B2CA}
    feeds: dict = field(default_factory=dict)         # klasse -> row
    bus: dict = field(default_factory=dict)           # bustype -> row
    texts: dict = field(default_factory=dict)
    locatiekoppen: list = field(default_factory=list)  # (pattern, locatie, sectie_fallback)

    @property
    def family(self):
        return "B2CA" if str(self.parameters.get("brandklasse", "")).upper() == "B2CA" else "CCA"

    def signal_cable(self, functietype):
        row = self.cable_by_type.get(functietype)
        if not row:
            return None
        return row["kabel_B2CA_JOBA"] if self.family == "B2CA" else row["kabel_CCA_DRAK"]

    def feed_cable(self, klasse):
        row = self.feeds.get(klasse)
        if not row:
            return None, "", ""
        kab = row["kabel_B2CA"] if self.family == "B2CA" else row["kabel_CCA"]
        ws_key = "ws_B2CA" if self.family == "B2CA" else "ws_CCA"
        ws = row.get(ws_key, "")
        if ws in ("", None):
            ws = row.get("ws", "")
        tpl_key = "sjabloon_B2CA" if self.family == "B2CA" else "sjabloon_CCA"
        tpl = row.get(tpl_key, "")
        if tpl in ("", None):
            tpl = row.get("bekabeling_naar_sjabloon", "")
        return kab, ws, tpl


def load_header_map(path):
    """Load an editable header-map workbook (config/header_map_<client>.xlsx).
    Returns {client_header_lowercase: canonical_field}. Sheet 'HeaderMap' with
    columns client_header / canonical_field. Blank rows ignored."""
    wb = load_workbook(path, data_only=True)
    ws = wb["HeaderMap"] if "HeaderMap" in wb.sheetnames else wb.active
    out = {}
    for r in _rows(ws):
        src = str(r.get("client_header", "")).strip().lower()
        dst = str(r.get("canonical_field", "")).strip()
        if src and dst:
            out[src] = dst
    return out


def load_config(path):
    wb = load_workbook(path, data_only=True)
    cfg = Config()
    for r in _rows(wb["0_Parameters"]):
        cfg.parameters[str(r["parameter"]).strip()] = r["waarde"]
    for r in _rows(wb["1_Synoniemen"]):
        cfg.synonyms.append((str(r[next(k for k in r if k.startswith("patroon"))]).strip().lower(),
                             str(r["functietype"]).strip(),
                             int(r["prioriteit"] or 0)))
    cfg.synonyms.sort(key=lambda t: (-t[2], -len(t[0])))
    for r in _rows(wb["2_Kabelkeuze"]):
        cfg.cable_by_type[str(r["functietype"]).strip()] = r
    for r in _rows(wb["3_Voedingen"]):
        cfg.feeds[str(r["klasse"]).strip()] = r
    for r in _rows(wb["4_Bus"]):
        cfg.bus[str(r["bustype"]).strip()] = r
    for r in _rows(wb["5_Vaste_teksten"]):
        cfg.texts[str(r["sleutel"]).strip()] = str(r["tekst"])
    if "6_Locatiekoppen" in wb.sheetnames:
        for r in _rows(wb["6_Locatiekoppen"]):
            pat = str(r[next(k for k in r if k.startswith("patroon"))]).strip().lower()
            if pat:
                cfg.locatiekoppen.append((pat, str(r.get("locatie", "")).strip(),
                                          str(r.get("sectie_fallback", "")).strip()))
        cfg.locatiekoppen.sort(key=lambda t: -len(t[0]))
    return cfg

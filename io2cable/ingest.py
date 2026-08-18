"""Step 1 — Ingest & normalization.

Layer A (here): deterministic parse of structured Excel function lists (functielijsten)
via a
per-client column mapping. Fully reproducible, no AI.

Layer B (outside this module): messy inputs (PDFs, scans, free text) are
normalized with AI assistance *into the same canonical CSV*, then loaded with
load_normalized(). The normalized table is the contract; it is ALWAYS written
out as normalized_review.xlsx for mandatory human sign-off before Step 2.
"""
import csv
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill
from .schema import NormRow, NORM_COLUMNS

# Client header text (lowercased, spaces ok) -> canonical NormRow field.
# Keys are what appears in a client's functielijst; values are the internal
# field names. Each canonical name is also accepted as a header in its own
# right, so it does not need to be repeated as an alias below.
_HEADER_ALIASES = {
    "omschrijving": ["beschrijving", "onderdeel"],
    "procescode":   ["proces code", "ref"],
    "fabricaat":    ["Fabrikaat"],
    "type":         [],
    "aantal":       [],
    "AI":           ["analoge ingang", "analoog in", "ai"],
    "AO":           ["analoge uitgang", "analoog uit", "au", "ao"],
    "DI":           ["digitale ingang", "digitaal in", "di"],
    "DO":           ["digitale uitgang", "digitaal uit", "do"],
    "DI_bedrijf":   ["bedrijfmelding", "bedrijf"],
    "DI_storing":   ["storingsmelding", "storing"],
    "DI_status":    ["statusmelding", "status"],
    "SOFT":         ["io bus-punt", "soft", "bus-punt"],
    "bus_protocol": ["data", "busprotocol", "protocol"],
    "bus_naam":     ["databus", "poort"],
    "voltage":      ["spanning", "voeding (v)", "v"],
    "power_kw":     ["vermogen", "vermogen (kw)", "kw"],
    "current_a":    ["stroom", "stroom (nom)", "a"],
    "va":           [],
    "opmerking":    ["specificatie"],
    "mr_flag":      ["m&r"],
}

DEFAULT_HEADER_MAP = {}
for _field, _aliases in _HEADER_ALIASES.items():
    DEFAULT_HEADER_MAP[_field.lower()] = _field
    for _alias in _aliases:
        DEFAULT_HEADER_MAP[_alias] = _field

_PARSE_ONLY = {"DI_bedrijf", "DI_storing", "DI_status", "mr_flag"}
_unknown = set(DEFAULT_HEADER_MAP.values()) - set(NORM_COLUMNS) - _PARSE_ONLY
assert not _unknown, f"header map targets unknown schema fields: {sorted(_unknown)}"

DERDEN_PAT = re.compile(
    r"derden|derde|vanuit hvk|vanuit e-verdeler|uit e-installatie|vanuit bmc", re.I)


ERR_TOKENS = ("#VALUE!", "#REF!", "#N/A", "#DIV", "#NAME")


def _num(v):
    """Parse an I/O count. '*1' style markers -> 1. Excel error tokens -> 0
    (the caller records a data-quality note). Returns (value, was_error)."""
    if v in (None, "", "-", "x", "X"):
        return 0
    s = str(v).strip()
    if s.startswith("*"):
        try:
            return int(s[1:] or 1)
        except ValueError:
            return 1
    if any(s.startswith(t[:2]) and t in s for t in ERR_TOKENS) or s.startswith("#"):
        return 0
    try:
        return int(float(s.replace(",", ".")))
    except ValueError:
        return 0


def _is_err(v):
    return isinstance(v, str) and v.strip().startswith("#")

_PROTOCOL_WORDS = ("modbus", "bacnet", "m-bus", "mbus", "knx", "lon", "profibus")


def _header_protocol(text):
    """Some layouts name the bus protocol in the COLUMN HEADER and put point
    counts in the cells -- e.g. a column headed 'modbus RTU' whose cells hold
    100, 10. Returns the header text when it names a protocol, else None.
    Evidence: project 7267 (Apparatuurlijst), 3 BMS cables lost without this."""
    s = str(text or "").strip()
    return s if any(w in s.lower() for w in _PROTOCOL_WORDS) else None

def _parse_electrical_spec(text):
    """Extract (voltage, current_a, power_kw) from free-text electrical specs, e.g.
    '230vac/6A' -> ('230', 6.0, None); '400V/10A/4kW' -> ('400', 10.0, 4.0);
    '24Vac/0-10VDC' -> ('24', None, None); 'EC,400V/2,2kW' -> ('400', None, 2.2).
    Dutch decimal commas handled. Voltage stays as a string (schema field is str),
    current_a and power_kw are floats (schema fields are Optional[float])."""
    if not text:
        return "", None, None
    t = str(text)
    voltage = ""
    current = power = None
    m = re.search(r"(\d{2,3})\s*v(?:ac|dc)?\b", t, re.I)
    if m:
        voltage = m.group(1)
    m = re.search(r"(\d+(?:[.,]\d+)?)\s*a\b(?!c)", t, re.I)
    if m:
        try:
            current = float(m.group(1).replace(",", "."))
        except ValueError:
            pass
    m = re.search(r"(\d+(?:[.,]\d+)?)\s*kw\b", t, re.I)
    if m:
        try:
            power = float(m.group(1).replace(",", "."))
        except ValueError:
            pass
    return voltage, current, power


def _fnum(v):
    if v in (None, "", "-"):
        return None
    try:
        return float(str(v).replace(",", "."))
    except ValueError:
        return None


def parse_excel(path, rk="RK?", header_map=None, sheet=None, mr_only=True):
    """Layer A: parse a structured function list. Detects the header row by scoring
    against the header map; carries group headers down; keeps provenance."""
    hmap = {**DEFAULT_HEADER_MAP, **(header_map or {})}
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    grid = list(ws.iter_rows(values_only=True))

    def score(row):
        return sum(1 for c in row if c and str(c).strip().lower() in hmap)
    hdr_i = max(range(min(len(grid), 30)), key=lambda i: score(grid[i]))
    if score(grid[hdr_i]) < 3:
        raise ValueError(f"No recognizable header row in {path}; supply a header_map.")
    cols, proto_cols = {}, {}
    for j, c in enumerate(grid[hdr_i]):
        key = str(c).strip().lower() if c else ""
        if key in hmap:
            cols.setdefault(hmap[key], j)
        elif _header_protocol(c):
            proto_cols[j] = _header_protocol(c)

    rows, group = [], ""
    for i, r in enumerate(grid[hdr_i + 1:], start=hdr_i + 2):
        def get(f):
            return r[cols[f]] if f in cols and cols[f] < len(r) else None
        desc = str(get("omschrijving") or "").strip()
        if not desc:
            continue
        if "mr_flag" in cols:
            if mr_only and str(get("mr_flag") or "").strip().lower() != "ja":
                # A non-M&R row is either a GROUP HEADER or an ACCESSORY (valve,
                # meetpunt, safety kit, TSA). Only a bare row -- no procescode, no
                # manufacturer, no type, no I/O -- is a header. Accessories carry
                # equipment data and must never become section names.
                # Evidence: Duitslandlaan/Fonkel emitted sections 'safety kit',
                # 'TSA PN16', 'Deelstroomfilter Deel-SEP GKW' from accessory rows.
                if (not get("procescode") and not get("fabricaat") and not get("type")
                        and not any(_num(get(f)) for f in ("AI", "AO", "DI", "DO"))):
                    group = desc
                continue
        else:
            # layout without M&R column (e.g. Coneco): a row with no quantity and
            # no I/O and no bus is a group header
            if not get("aantal") and not get("procescode") and not get("bus_protocol") and not any(_num(get(f)) for f in ("AI", "AO", "DI", "DO", "SOFT")):
                group = desc
                continue
        bad = [f for f in ("AI", "AO", "DI", "DO", "SOFT") if _is_err(get(f))]
        di = _num(get("DI")) + _num(get("DI_bedrijf")) + _num(get("DI_storing")) + _num(get("DI_status"))
        row = NormRow(
            rk=rk, group=group, procescode=str(get("procescode") or "").strip(),
            omschrijving=desc, fabricaat=str(get("fabricaat") or "").strip(),
            type=str(get("type") or "").strip(), aantal=_num(get("aantal")) or 1,
            AI=_num(get("AI")), AO=_num(get("AO")), DI=di, DO=_num(get("DO")),
            SOFT=_num(get("SOFT")),
            voltage=str(get("voltage") or "").strip(),
            power_kw=_fnum(get("power_kw")), current_a=_fnum(get("current_a")),
            va=_fnum(get("va")),
            bus_protocol=str(get("bus_protocol") or "").strip(),
            bus_naam=str(get("bus_naam") or "").strip(),
            opmerking=str(get("opmerking") or "").strip(),
            source_ref=f"{path}:{ws.title}:r{i}",
        )
        row.regelkast_spec = str(get("regelkast_spec") or "").strip()
        if not row.bus_protocol and proto_cols:
            # Protocol named in the column header (see _header_protocol): a
            # non-empty cell in that column means the device is on that bus.
            for _j, _proto in proto_cols.items():
                if _j < len(r) and _num(r[_j]):
                    row.bus_protocol = _proto
                    break
        if not row.voltage and not row.power_kw and not row.current_a:
            # Layouts with no dedicated U/I/P columns (Append1) embed the electrical
            # spec as free text in Opmerking: '230vac/6A', '400V/10A/4kW',
            # '24Vac/0-10VDC', 'EC,400V/2,2kW'. Parse it from THERE ONLY -- never
            # from 'toevoeging', which can coincidentally contain a bare voltage
            # ('Spanning aanwezig' -> toevoeging='230V') describing a monitored
            # rail, not a device's own feed requirement.
            v, i_, p = _parse_electrical_spec(row.opmerking)
            if v: row.voltage = v
            if i_: row.current_a = i_
            if p: row.power_kw = p
        if not row.voltage and row.regelkast_spec.lower().startswith("voeding"):
            # Coneco layouts put the electrical spec in the panel column and leave U
            # empty ('Voeding pomp 230V max 1kW'); the spec string IS the specification.
            # Evidence: Tilburg PR20267273 -- 14/24 feed rows have voltage ONLY here.
            m = re.search(r"(\d{3})\s*v", row.regelkast_spec.lower())
            if m and m.group(1) in ("230", "400"):
                row.voltage = m.group(1)
        if bad:
            row.opmerking = (row.opmerking + f" [!formula errors in {','.join(bad)} — correct in review]").strip()
        row.derden_flag = bool(DERDEN_PAT.search(row.opmerking + " " + desc))
        if re.search(r"dak|bovendaks", f"{group} {desc} {row.opmerking}", re.I):
            row.locatie = "op dak"
        rows.append(row)
    return rows


def load_normalized(path):
    """Load the canonical CSV (Layer B output or a reviewed/corrected table)."""
    out = []
    with open(path, newline="", encoding="utf-8") as f:
        for d in csv.DictReader(f):
            row = NormRow()
            for k, v in d.items():
                if k not in NORM_COLUMNS or v is None:
                    continue
                cur = getattr(row, k)
                if isinstance(cur, bool):
                    setattr(row, k, str(v).strip().lower() in ("1", "true", "ja", "yes"))
                elif isinstance(cur, int):
                    setattr(row, k, _num(v))
                elif k in ("power_kw", "current_a", "va"):
                    setattr(row, k, _fnum(v))
                else:
                    setattr(row, k, str(v).strip())
            out.append(row)
    return out


def write_review(rows, path):
    """Mandatory human-review file: the full normalized table, one row per input row."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Normalized"
    ws.append(NORM_COLUMNS)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        c.fill = PatternFill("solid", start_color="1F4E78")
    for r in rows:
        d = r.as_dict()
        ws.append([d[c] for c in NORM_COLUMNS])
    for col, w in zip("ABCDEFG", (8, 22, 12, 44, 14, 30, 7)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    wb.save(path)

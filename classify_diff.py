"""Classify-diff — show how a change to the match text or to 1_Synoniemen
changes the PRIMARY functietype of every row, across every project.

Self-contained on purpose: it reads 1_Synoniemen straight from the config
workbook and re-implements the primary-match rule from classify.py, so it has
no import coupling to the package and can be run before/after an edit without
touching the pipeline.

    python classify_diff.py --config config\\kabelconfig.xlsx ^
        --rows projects\\duitslandlaan\\normalized.csv ^
        --rows projects\\fonkel\\normalized.csv ^
        --rows inputs\\7222\\out\\normalized_review.xlsx

Accepts normalized.csv fixtures and normalized_review.xlsx files (both carry
the canonical NORM_COLUMNS). Prints, per project, every row whose primary
functietype differs between the current match text and the proposed one.

IMPORTANT: keep MATCH_FIELDS_BEFORE in sync with the live classify.py. If
classify.py has changed, this tool's "before" is wrong and so is the diff.
"""
import argparse
import csv
import os
from collections import Counter

from openpyxl import load_workbook

# --- the two match texts being compared -------------------------------------
MATCH_FIELDS_BEFORE = ("omschrijving", "type", "opmerking")
MATCH_FIELDS_AFTER = ("procescode", "omschrijving", "type", "opmerking")


def load_synonyms(config_path):
    """Read 1_Synoniemen -> [(patroon, functietype, prioriteit), ...]."""
    wb = load_workbook(config_path, data_only=True)
    ws = wb["1_Synoniemen"]
    rows, header_seen = [], False
    for r in ws.iter_rows(values_only=True):
        cells = [c for c in r if c not in (None, "")]
        if len(cells) < 3:
            continue
        pat, ftype, prio = r[0], r[1], r[2]
        if not header_seen:
            # skip the title/description block and the header row itself
            if isinstance(prio, (int, float)):
                header_seen = True
            else:
                continue
        if not isinstance(prio, (int, float)):
            continue
        rows.append((str(pat).strip().lower(), str(ftype).strip(), int(prio)))
    if not rows:
        raise SystemExit("No synonym rows parsed - check the 1_Synoniemen layout.")
    return rows


def primary(text, synonyms):
    """Primary functietype only: highest priority, then longest pattern.
    Mirrors classify.py's ordering. Does NOT reproduce the DERDEN/REGELKAST
    demotion or the I/O-signature fallbacks - this tool compares dictionary
    hits, which is where a match-text change has its effect."""
    hits = [(p, len(pat), f) for pat, f, p in synonyms if pat in text]
    if not hits:
        return None
    hits.sort(key=lambda h: (-h[0], -h[1]))
    return hits[0][2]


def read_rows(path):
    """Yield dicts from a normalized.csv or a normalized_review.xlsx."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        with open(path, newline="", encoding="utf-8") as f:
            for d in csv.DictReader(f):
                yield d
    else:
        ws = load_workbook(path, data_only=True).active
        grid = list(ws.iter_rows(values_only=True))
        header = [str(c).strip() if c else "" for c in grid[0]]
        for r in grid[1:]:
            yield {h: ("" if v is None else str(v)) for h, v in zip(header, r)}


def text_of(row, fields):
    return " ".join(str(row.get(f, "") or "") for f in fields).lower()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", required=True)
    ap.add_argument("--rows", action="append", required=True,
                    help="normalized.csv or normalized_review.xlsx (repeatable)")
    ap.add_argument("--show", type=int, default=15,
                    help="max example rows printed per project")
    args = ap.parse_args()

    syn = load_synonyms(args.config)
    print(f"{len(syn)} synonym patterns loaded from {args.config}")
    print(f"before: {' + '.join(MATCH_FIELDS_BEFORE)}")
    print(f"after : {' + '.join(MATCH_FIELDS_AFTER)}\n")

    grand = Counter()
    for path in args.rows:
        name = os.path.basename(os.path.dirname(path)) or path
        changed, total, examples = 0, 0, []
        transitions = Counter()
        for row in read_rows(path):
            if not (row.get("omschrijving") or "").strip():
                continue
            total += 1
            b = primary(text_of(row, MATCH_FIELDS_BEFORE), syn)
            a = primary(text_of(row, MATCH_FIELDS_AFTER), syn)
            if b != a:
                changed += 1
                transitions[(b, a)] += 1
                if len(examples) < args.show:
                    examples.append((row.get("procescode", ""),
                                     (row.get("omschrijving", "") or "")[:44], b, a))
        verdict = "NO CHANGE" if changed == 0 else f"{changed} CHANGED"
        print(f"=== {name}: {total} rows -> {verdict}")
        for (b, a), n in transitions.most_common():
            print(f"      {n:5d}  {str(b):26s} -> {a}")
        for proc, ond, b, a in examples:
            print(f"        {str(proc)[:24]:26s} | {ond:46s} {b} -> {a}")
        grand[name] = changed
        print()

    print("summary:", dict(grand))
    print("\nA project showing NO CHANGE is unaffected by the edit.")
    print("Any project showing changes needs each transition judged against "
          "its manual before the edit is merged.")


if __name__ == "__main__":
    main()
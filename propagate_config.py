import shutil
from openpyxl import load_workbook

for name in ['boerhaave', 'duitslandlaan', 'fonkel']:
    p = rf'config\kabelconfig_{name}.xlsx'
    old = load_workbook(p)['0_Parameters']
    params = [[c.value for c in row] for row in old.iter_rows()]
    shutil.copy(r'config\kabelconfig_update.xlsx', p)
    wb = load_workbook(p)
    ws = wb['0_Parameters']
    for r, row in enumerate(params, 1):
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=v)
    wb.save(p)
    print('rebuilt', p)